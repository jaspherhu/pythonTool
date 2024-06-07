# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
import os

tips = '请选择Hr发布的考勤xlsx文件'

def update_tips(user_tips): 
    tips = 'Tips:' + user_tips
    tip.configure(text=tips)  # 更改 Label 文本

def open_file():
    file_path = filedialog.askopenfilename()  # 打开文件选择窗口
    select_path.set(file_path)
    #print(file_path)  # 输出选择的文件路径
    if (file_path):
        update_tips("Hr文件选择完成...") 
    else:
        update_tips("Hr文件选择异常!!!") 

def out_file():
    file_path = filedialog.askopenfilename()  # 打开文件选择窗口
    output_path.set(file_path)
    print(file_path)  # 输出选择的文件路径

def select_folder():
    # 文件夹选择
    file_path = filedialog.askdirectory()  # 使用askdirectory函数选择文件夹
    output_path.set(file_path)
    #print(file_path)  # 输出选择的文件路径
    if (file_path):
        update_tips("文件夹选择完成...") 
    else:
        update_tips("文件夹选择异常!!!") 

def contains_char(s, char):
    return char in s

def run_app():
    times = '加班时长'
    t_name = '员工姓名'
    t_short = '缺卡次数'
    after_20 = "加班到20点后次数"
    sunday_ = "周末加班天数"
    
    idx = 0
    name = [0]   #没有初始化无法直接访问,会报告溢出
    time_total = [0]

    item = {
        'id': 0,
        'name': '',
        'time': 0,
        'short': 0,
        'after_20': 0,
        'wekend': 0
    }

    list = [0,0,0,0]
    #ite = [0,t_name,times,t_short]
    ite = []

    update_tips("分析源文件中...") 
    file_home = select_path.get()
    file_out = output_path.get() + "/sin_kaoqin.xlsx"
    #file_home = 'd:/gitee/py_tool/02 requests/盛弘电气4月考勤数据-充换电服务BU研发部.xlsx'
    wb = load_workbook(filename=file_home) # 如果不打开本地文件的话,创建一个内存Workbook()
    ws = wb.active
    web_s1 = wb['打卡明细']
    #print(web_s1['A2'].value)
    #第一行的姓名当成了缓冲,
    
    for line in range(1,web_s1.max_row+1):
        if(web_s1.cell(column=2, row=line).value != t_name):
            #name.append(t_name)     #直接使用角标访问会提示溢出
            #time_total.append(times)
            #print(name[idx],time_total[idx])
            item['id'] = idx
            item['name'] = t_name
            item['time'] = times
            item['short'] = t_short
            item['after_20'] = after_20
            item['wekend'] = sunday_
            L = item.copy() #深浅拷贝导致
            ite.append(L)  #extend #append

            #list[0] = idx
            #list[1] = t_name
            #list[2] = times
            #list[3] = t_short
            #L = list.copy() #深浅拷贝导致,为了节省内存干这种事情可耻 = https://www.cnblogs.com/Xloading/p/15226696.html 
            #ite.append(L)  #extend #append
            #print(items)
            
            times = 0
            t_short = 0
            sunday_ = 0
            after_20 = 0
            idx = idx + 1
            t_name = web_s1.cell(column=2, row=line).value
            #print(ite)

        str = web_s1.cell(column=8, row=line).value
        str2 = web_s1.cell(column=5, row=line).value
        if(str == None):
            continue

        if(str == '--'):
            continue

        if(line == 1):
            continue
        
        if((contains_char(str2, '周末')) and (0 == contains_char(str2, '补班'))):
            if(contains_char(str, ':')):
                sunday_ = sunday_ + 1
            continue

        if((contains_char(str, '次日'))):
            p = str.split('次日')
            str = p[1]

        if(contains_char(str, ':')):
            t = str.split(':')         
            times = times + int(t[0])*60 + int(t[1]) - 18*60

            if(int(t[0]) >= 20 ):
                after_20 = after_20 + 1
        else:
            if(contains_char(str, '未打卡')):
                t_short = t_short + 1
                #print(str)
            else:
                print('err_act',str)

    update_tips("准备生成文件中...") 
    web_s2 = wb.create_sheet("统计结果")#创建mysheet表 
    web_s2.title = "统计结果" #创建完了后还可以改名字
    web_s3 = wb['统计结果'] #sheet创建完了后就可以对他进行编辑了 s3
    #print(ite)
    print("total_people_in_excel:",idx)

    d = 1
    #想办法大到小排序
    #for d in range(1,idx):
    for item in ite:
        #web_s3.cell(row=d, column=1).value = name[d]
        #web_s3.cell(row=d, column=2).value = time_total[d]
        #print(name[d],time_total[d])

        web_s3.cell(row=d, column=1).value = item['id']
        web_s3.cell(row=d, column=2).value = item['name']
        web_s3.cell(row=d, column=3).value = item['time']
        web_s3.cell(row=d, column=4).value = item['short']
        web_s3.cell(row=d, column=5).value = item['after_20']
        web_s3.cell(row=d, column=6).value = item['wekend']

        #web_s3.cell(row=d, column=1).value = item[0]
        #web_s3.cell(row=d, column=2).value = item[1]
        #web_s3.cell(row=d, column=3).value = item[2]
        #web_s3.cell(row=d, column=4).value = item[3]
        d = d + 1
        #print('d:',d,it)
    #wb.save('d:/gitee/py_tool/02 requests/work_time_copy.xlsx') #file_out
    wb.save(file_out) #file_out
    #print(file_out)  # 输出选择的文件路径
    update_tips("文件生成完成,点击按钮打开查看...") 

def read_file():
    update_tips("为您打开文件中...") 
    read_path = output_path.get() + "/sin_kaoqin.xlsx"
    if (output_path.get()):
        # 使用系统命令打开选择的Excel文件
        os.startfile(read_path)
        print(read_path)  # 输出选择的文件路径
    else:
        #user_tips = 
        update_tips("您还没生成目标文件呢~")  
        #print("open file err:",read_path)  # 输出选择的文件路径
        #print(tips)


def exit_program():
    root.destroy()

root = tk.Tk()
root.title("盛弘考勤统计小助手...")
root.geometry('400x200')


# 初始化Entry控件的textvariable属性值
select_path = tk.StringVar()
output_path = tk.StringVar()
# 布局控件
tk.Label(root, text="公司考勤原文件路径：").grid(column=0, row=0, rowspan=3)
tk.Entry(root, textvariable = select_path).grid(column=1, row=0, rowspan=3)

tk.Label(root, text="汇总后文件路径：").grid(column=0, row=3, rowspan=3)
tk.Entry(root, textvariable = output_path).grid(column=1, row=3, rowspan=3)

tip = tk.Label(root, text='Tips:' + tips, fg='red')#.grid(column=0, row=10, rowspan=3)
#tip = tk.Label(root, textvariable='tips:' + tips, fg='red')
tip.place(y=80, x=25)

tk.Button(root, text="上传HR发布打卡表格", command=open_file).grid(row=0, column=2)
tk.Button(root, text="统计后生成存放文件夹", command=select_folder).grid(row=3, column=2)

button1 = tk.Button(root, text="点击运行", command=run_app)#.grid(row=20, column=0)
button1.place(y=125, x=25)
button2 = tk.Button(root, text="打开刚生成的文件", command=read_file)#.grid(row=20, column=1)
button2.place(y=125, x=135)
button3 = tk.Button(root, text="退出程序", command=exit_program)#.grid(row=20, column=2)
button3.place(y=125, x=305)
root.mainloop()
